from openpyxl import load_workbook

class ExcelFunctions:
    def __init__(self, file_name, sheet_name):
        self.file = file_name
        self.sheet = sheet_name
    
    # fetch the row count
    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row
    
    # fetch the column count
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column
    
    # read the data from excel file
    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row = row_number, column= column_number).value
    
    # write the data to the excel file
    def write_data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row= row_number, column= column_number).value = data
        workbook.save(self.file)
    